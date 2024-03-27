import unittest
import sys
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)

from app.routes.omegaTron.strategy_route import update_kill_switch

class TestKillSwitchFunction(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            cls.workbook = load_workbook("kill_switch.xlsx")
        except FileNotFoundError:
            cls.workbook = Workbook()

        cls.sheet = cls.workbook.active
        cls.sheet.title = "Login Results"
        if cls.sheet.max_row == 1:
            cls.sheet.append(
                [
                    "test case",
                    "result",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(cls):
        cls.workbook.save("kill_switch.xlsx")

    def write_to_excel(self, test_case, test_result):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, date, time])

    def test_kill_switch_toggle(self):
        try:
            initial_state = update_kill_switch()
            final_state = update_kill_switch()
            self.assertNotEqual(final_state, initial_state)
            final_state = update_kill_switch()
            self.assertEqual(final_state, initial_state)
            self.write_to_excel("test_kill_switch_toggle", "Passed")
        except Exception as e:
            print(format(str(e)))
            self.write_to_excel("test_kill_switch_toggle", "Failed")

if __name__ == '__main__':
    unittest.main()
