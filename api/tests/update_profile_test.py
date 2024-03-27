import unittest
from unittest.mock import patch
from flask import json
import os, sys
from openpyxl import Workbook, load_workbook
from datetime import datetime

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)
from app import app

class TestUserProfileUpdate(unittest.TestCase):

    @classmethod
    def setUpClass(self):
        self.app = app.test_client()
        self.app.testing = True
        try:
            self.workbook = load_workbook("update_profile.xlsx")
        except FileNotFoundError:
            self.workbook = Workbook()

        self.sheet = self.workbook.active
        self.sheet.title = "Test result"
        if self.sheet.max_row == 1:
            self.sheet.append(
                [
                    "test case",
                    "result",
                    "notes",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(self):
        self.workbook.save("update_profile.xlsx")

    def write_to_excel(self, test_case, test_result, notes):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, notes, date, time])

    def test_invalid_email_id(self):
        try:
            login_result = self.app.post('/login', json={"email": "new_user@xorlabs.com","password": "securePassword123"})
            token = json.loads(login_result.data)['token']
            response = self.app.post('/update_profile', json={'name': 'test user', 'email': 'new_user@xorlabs.com','dob' : '1990-05-15', 'mobile_number' : '+91xxxxxxxxxx'}, headers={'Authorization': f'Bearer {token}'})
            data = response.get_json()
            self.assertEqual(response.status_code, 409)
            self.assertEqual(data['message'], 'Email id already exists')
            self.write_to_excel('test_invalid_email_id','Success','')
        except Exception as e:
            self.write_to_excel('test_invalid_email_id','Failure',str(e))

    def test_invalid_token(self):
        try:
        # Invalid token
            response = self.app.post('/update_profile', json={'name': 'John Doe', 'email':  'new_user@xorlabs.com' ,'dob' : '1990-05-15', 'mobile_number' : '+91xxxxxxxxxx'}, headers={'Authorization': 'Bearer invalid_token'})
            data = response.get_json()
            self.assertEqual(response.status_code, 401)
            self.assertIn('message', data)
            self.write_to_excel('test_invalid_token','Success','')
        except Exception as e:
            self.write_to_excel('test_invalid_token','Failure',str(e))

    def test_missing_fields(self):
        try:
            login_result = self.app.post('/login', json={"email": "new_user@xorlabs.com","password": "securePassword123"})
            token = json.loads(login_result.data)['token']
            # Missing email field
            response = self.app.post('/update_profile', json={'name': 'John Doe'}, headers={'Authorization': f'Bearer {token}'})
            data = response.get_json()
            self.assertEqual(response.status_code, 400)
            self.assertIn('message', data)
            self.write_to_excel('test_missing_fields','Success','')
        except Exception as e:
            self.write_to_excel('test_missing_fields','Failure',str(e))



if __name__ == '__main__':
    unittest.main()
