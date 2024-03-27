import unittest
from unittest.mock import patch
from datetime import datetime
import json
import os
import sys
from openpyxl import Workbook, load_workbook

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)

from app import app

class TestPremiumStrategy(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            cls.workbook = load_workbook("delta_tron_test.xlsx")
        except FileNotFoundError:
            cls.workbook = Workbook()

        cls.sheet = cls.workbook.active
        cls.sheet.title = "Delta tron Results"
        if cls.sheet.max_row == 1:
            cls.sheet.append(
                [
                    "test case",
                    "result",
                    "notes",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(cls):
        cls.workbook.save("delta_tron_test.xlsx")

    def write_to_excel(self, test_case, test_result, notes):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, notes, date, time])

    @patch('app.routes.deltaTron.premium_route.is_strategy_running', True)
    def test_premium_strategy_with_running_strategy(self):
        try:
        # Create a test client
            app.testing = True
            self.app = app.test_client()

            # Prepare mock data
            mock_request_data = {
                "entry_price": "[10.0, 20.0, 30.0]",
                "usertype": "mock_user"
            }

            # Make a POST request to the route
            response = self.app.post('/deltaTron_strategy', data=json.dumps(mock_request_data), content_type='application/json')

            expected_response_content = "Found another instance of scheduler running, skipping this schedule.\n"
            self.assertEqual(response.data.decode('utf-8'), expected_response_content)
            self.write_to_excel("test_premium_strategy_with_running_strategy", "Passed", "")
        except Exception as e:
            self.write_to_excel("test_premium_strategy_with_running_strategy", "Failed", str(e))


    @patch('app.routes.deltaTron.premium_route.is_market_closing', return_value=True)
    def test_premium_strategy_with_market_closing(self, mock_is_market_closing):
        try:
            app.testing = True
            self.app = app.test_client()
            # Mock the request data
            test_premium_strategy_with_market_closing = {
                "entry_price": "[10.0, 20.0, 30.0]",
                "usertype": "mock_user"
            }

            # Call the premium_strategy function
            result = self.app.post('/deltaTron_strategy', data=json.dumps(test_premium_strategy_with_market_closing), content_type='application/json')
            result_string = result.data.decode('utf-8')

            # Assert that the function returns the expected result
            expected_result = "All active orders have been bought back"
            self.assertEqual(result_string, expected_result)
            self.write_to_excel("test_premium_strategy_with_market_closing", "Passed", "")
        except Exception as e:
            self.write_to_excel("test_premium_strategy_with_market_closing", "Failed", str(e))

    # @patch('app.routes.deltaTron.premium_route.lastStep', None)
    def test_wrong_format_data(self):
        try:
        # Mock the request data
            app.testing = True
            self.app = app.test_client()
            # Mock the request data
            test_premium_strategy_with_market_closing = {
                "entry_price": [10.0, 20.0, 30.0],
                "usertype": "mock_user"
            }

            # Call the premium_strategy function
            response = self.app.post('/deltaTron_strategy', data=json.dumps(test_premium_strategy_with_market_closing), content_type='application/json')

            # Assert that the function raises an exception
            self.assertEqual(response.status_code, 500)
            self.assertIn('error', json.loads(response.data))
            self.write_to_excel("test_wrong_format_data", "Passed", "")
        except Exception as e:
            self.write_to_excel("test_wrong_format_data", "Failed", str(e))


if __name__ == '__main__':
    unittest.main()
