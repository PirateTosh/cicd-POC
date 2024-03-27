import unittest
from unittest.mock import patch
from flask import json
import os,sys
from openpyxl import Workbook, load_workbook
from datetime import datetime

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)
from app import app

class TestOTPVerification(unittest.TestCase):

    @classmethod
    def setUpClass(self):
        self.app = app.test_client()
        self.app.testing = True
        try:
            self.workbook = load_workbook("user_Verification.xlsx")
        except FileNotFoundError:
            self.workbook = Workbook()

        self.sheet = self.workbook.active
        self.sheet.title = "Verification result"
        if self.sheet.max_row == 1:
            self.sheet.append(
                [
                    "test case",
                    "result",
                    "notes",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(self):
        self.workbook.save("user_Verification.xlsx")

    def write_to_excel(self, test_case, test_result, notes):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, notes, date, time])

    @patch('app.routes.userModule.user_verification.send_otp_email')
    def test_send_otp_success(self, mock_send_email):
        try:
            mock_send_email.return_value = True
            response = self.app.post('/send_otp', json={'email': 'test@example.com'})
            data = json.loads(response.data.decode())
            self.assertEqual(response.status_code, 200)
            self.assertEqual(data['message'], 'OTP sent successfully')
            self.write_to_excel('test_send_otp_success','Success','')
        except Exception as e:
            self.write_to_excel('test_send_otp_success','Failure',str(e))

    @patch('app.routes.userModule.user_verification.send_otp_email')
    def test_send_otp_failure(self, mock_send_email):
        try:
            mock_send_email.return_value = False
            response = self.app.post('/send_otp', json={'email': 'test@example.com'})
            data = json.loads(response.data.decode())
            self.assertEqual(response.status_code, 500)
            self.assertEqual(data['error'], 'Failed to send OTP')
            self.write_to_excel('test_send_otp_failure','Success','')
        except Exception as e:
            self.write_to_excel('test_send_otp_failure','Failure',str(e))

    @patch('app.routes.userModule.user_verification.otp_store', {'test@example.com': '123456'})
    def test_check_otp_success(self):
        try:
            response = self.app.post('/check_otp', json={'email': 'test@example.com', 'otp': '123456'})
            data = json.loads(response.data.decode())
            self.assertEqual(data['message'], 'OTP verification successful')
            self.write_to_excel('test_check_otp_success','Success','')
        except Exception as e:
            print(e)
            self.write_to_excel('test_check_otp_success','Failed', str(e))

    def test_check_otp_failure_invalid_email(self):
        try:
            response = self.app.post('/check_otp', json={'email': '', 'otp': '123456'})
            data = json.loads(response.data.decode())
            self.assertEqual(data['message'], 'Invalid email format')
            self.write_to_excel('test_check_otp_failure_invalid_email','Success','')
        except Exception as e:
            print(e)
            self.write_to_excel('test_check_otp_failure_invalid_email','Failed',str(e))

    @patch('app.routes.userModule.user_verification.otp_store', {'test@example.com': '123456'})
    def test_check_otp_failure_invalid_otp(self):
        try:
            response = self.app.post('/check_otp', json={'email': 'test@example.com', 'otp': '654321'})
            data = json.loads(response.data.decode())
            self.assertEqual(data['error'], 'Invalid OTP')
            self.write_to_excel('test_check_otp_failure_invalid_otp','Success','')
        except Exception as e:
            print(e)
            self.write_to_excel('test_check_otp_failure_invalid_otp','Failed',str(e))

if __name__ == '__main__':
    unittest.main()
