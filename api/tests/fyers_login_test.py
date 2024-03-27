import unittest
from unittest.mock import patch, MagicMock
from datetime import datetime
import sys
import os
from openpyxl import Workbook, load_workbook

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)
from app.utils.fyers_auth_service import generate_fyers_instance
from app.config import SUCCESS

FYERS_ID = "XV30279"
APP_ID = "2"  # Value 2 is used for web login
CLIENT_ID = "9HEI4IDQEK-100"
APP_SECRET = "494K0HOB2C"
REDIRECT_URI = "https://trade.fyers.in/api-login/redirect-uri/index.html"
TOTP_KEY = "YWLD6PGLM3JCMW2QBBJ3AWUKCMVQT3U3"
# PIN = 9999
APP_ID_HASH = "0273fdb25e81e790585f43c0b5fe2a3d332988086bfe45f7850e2d9629dd52d0"


class TestSuccessfulQuoteRetrieval(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            cls.workbook = load_workbook("fyers_login_test.xlsx")
        except FileNotFoundError:
            cls.workbook = Workbook()
        cls.sheet = cls.workbook.active
        cls.sheet.title = "Login Results"
        if cls.sheet.max_row == 1:
            cls.sheet.append(
                [
                    "test_case",
                    "Login_result",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(cls):
        cls.workbook.save("fyers_login_test.xlsx")

    def write_to_excel(self, test_case, test_result):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, date, time])

    # ? Function with correct details
    @patch("app.utils.fyers_auth_service.fyers_instance", None)
    @patch("app.utils.fyers_auth_service.send_login_otp")
    @patch("app.utils.fyers_auth_service.verify_otp")
    @patch("app.utils.fyers_auth_service.verify_pin")
    @patch("app.utils.fyers_auth_service.get_auth_code")
    @patch("app.utils.fyers_auth_service.fyersModel.SessionModel")
    @patch("app.utils.fyers_auth_service.fyersModel.FyersModel")
    def test_with_correct_access_token(
        self,
        mock_FyersModel,
        mock_SessionModel,
        mock_get_auth_code,
        mock_verify_pin,
        mock_verify_otp,
        mock_send_login_otp,
    ):
        mock_send_login_otp.return_value = (SUCCESS, "request_key")
        mock_verify_otp.return_value = (SUCCESS, "verify_totp_request_key")
        mock_verify_pin.return_value = (SUCCESS, "access_token")
        mock_get_auth_code.return_value = (SUCCESS, "auth_code")
        mock_session_instance = MagicMock()
        mock_SessionModel.return_value = mock_session_instance
        mock_response = {
            "access_token": "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJpc3MiOiJhcGkubG9naW4uZnllcnMuaW4iLCJpYXQiOjE3MDc4MDc3NjAsImV4cCI6MTcwNzgzNzc2MCwibmJmIjoxNzA3ODA3MTYwLCJhdWQiOiJbXCJ4OjBcIiwgXCJ4OjFcIiwgXCJ4OjJcIiwgXCJkOjFcIiwgXCJkOjJcIiwgXCJ4OjFcIiwgXCJ4OjBcIl0iLCJzdWIiOiJhdXRoX2NvZGUiLCJkaXNwbGF5X25hbWUiOiJYVjMwMjc5Iiwib21zIjoiSzEiLCJoc21fa2V5IjoiMjdlMjgxM2ZiZjg4Mjg1Mzg3NmZmZmViZTI1OWI4MWUzNTIwMWRkNTdmNWJkNGJmMTA1YzRjM2IiLCJub25jZSI6IiIsImFwcF9pZCI6IjlIRUk0SURRRUsiLCJ1dWlkIjoiNWZjOGI3MGRiMDFmNDM4YTg3ZTY1MTMxM2M5NGYxYzEiLCJpcEFkZHIiOiIwLjAuMC4wIiwic2NvcGUiOiIifQ.S2lNpEZiPUN-GRU2chLvoi31kuHpThBJdji-pP2C4AA"
        }
        mock_session_instance.generate_token.return_value = mock_response
        result, status_message = generate_fyers_instance()
        self.write_to_excel("test_with_correct_access_token", status_message)

    #! Function with fake access details
    @patch("app.utils.fyers_auth_service.fyers_instance", None)
    @patch("app.utils.fyers_auth_service.send_login_otp")
    @patch("app.utils.fyers_auth_service.verify_otp")
    @patch("app.utils.fyers_auth_service.verify_pin")
    @patch("app.utils.fyers_auth_service.get_auth_code")
    @patch("app.utils.fyers_auth_service.fyersModel.SessionModel")
    @patch("app.utils.fyers_auth_service.fyersModel.FyersModel")
    def test_with_fake_access_token(
        self,
        mock_FyersModel,
        mock_SessionModel,
        mock_get_auth_code,
        mock_verify_pin,
        mock_verify_otp,
        mock_send_login_otp,
    ):
        mock_send_login_otp.return_value = (SUCCESS, "request_key")
        mock_verify_otp.return_value = (SUCCESS, "verify_totp_request_key")
        mock_verify_pin.return_value = (SUCCESS, "access_token")
        mock_get_auth_code.return_value = (SUCCESS, "auth_code")
        mock_session_instance = MagicMock()
        mock_SessionModel.return_value = mock_session_instance
        mock_response = {"access_token": "access_token"}
        mock_session_instance.generate_token.return_value = mock_response
        result, status_message = generate_fyers_instance()
        self.write_to_excel("test_with_no_access_token", status_message)

    #! Function with no access token
    @patch("app.utils.fyers_auth_service.fyers_instance", None)
    @patch("app.utils.fyers_auth_service.send_login_otp")
    @patch("app.utils.fyers_auth_service.verify_otp")
    @patch("app.utils.fyers_auth_service.verify_pin")
    @patch("app.utils.fyers_auth_service.get_auth_code")
    @patch("app.utils.fyers_auth_service.fyersModel.SessionModel")
    def test_with_no_access_token(
        self,
        mock_SessionModel,
        mock_get_auth_code,
        mock_verify_pin,
        mock_verify_otp,
        mock_send_login_otp,
    ):
        mock_send_login_otp.return_value = (SUCCESS, "request_key")
        mock_verify_otp.return_value = (SUCCESS, "verify_totp_request_key")
        mock_verify_pin.return_value = (SUCCESS, "access_token")
        mock_get_auth_code.return_value = (SUCCESS, "auth_code")
        mock_session_instance = MagicMock()
        mock_SessionModel.return_value = mock_session_instance
        mock_response = {"access_token": ""}
        mock_session_instance.generate_token.return_value = mock_response
        result, status_message = generate_fyers_instance()
        self.write_to_excel("test_with_no_access_token", status_message)


if __name__ == "__main__":
    unittest.main()
