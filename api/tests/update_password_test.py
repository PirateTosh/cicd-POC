import unittest
from unittest.mock import patch
from flask import json
import os , sys
from openpyxl import Workbook, load_workbook
from datetime import datetime

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)
from app import app


class UpdatePasswordTestCase(unittest.TestCase):
    @classmethod
    def setUpClass(self):
        self.app = app.test_client()
        self.app.testing = True
        try:
            self.workbook = load_workbook("update_password.xlsx")
        except FileNotFoundError:
            self.workbook = Workbook()

        self.sheet = self.workbook.active
        self.sheet.title = "Test result"
        if self.sheet.max_row == 1:
            self.sheet.append(
                [
                    "test case",
                    "result",
                    "notes",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(self):
        self.workbook.save("update_password.xlsx")

    def write_to_excel(self, test_case, test_result, notes):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, notes, date, time])

    @patch('app.routes.userModule.update_password.execute_query')
    def test_update_password_success(self, mock_execute_query):
        try:
            """Test updating password successfully"""
            mock_execute_query.side_effect = [ # Mock the execution result of the SELECT and UPDATE queries
                [{'password': 'hashed_password'}],  # Mock result for SELECT query
                True  # Mock result for UPDATE query
            ]

            response = self.app.post('/update_password', data=json.dumps({
                'email': 'test@example.com',
                'password': 'new_password'
            }), content_type='application/json')
            
            self.assertEqual(response.status_code, 200)
            self.assertEqual(json.loads(response.data)['message'], 'Password updated successfully')
            self.write_to_excel('test_update_password_success','Success','')
        except Exception as e:
            self.write_to_excel('test_update_password_success','Failure',str(e))

    @patch('app.routes.userModule.update_password.execute_query')
    def test_update_password_email_not_found(self, mock_execute_query):
        try:
            """Test updating password when email is not found"""
            mock_execute_query.return_value = []  # Simulate email not found in database

            response = self.app.post('/update_password', data=json.dumps({
                'email': 'nonexistent@example.com',
                'password': 'new_password'
            }), content_type='application/json')
            
            self.assertEqual(response.status_code, 409)
            self.assertEqual(json.loads(response.data)['message'], 'Email not found')
            self.write_to_excel('test_update_password_email_not_found','Success','')
        except Exception as e:
            self.write_to_excel('test_update_password_email_not_found','Failure',str(e))

    def test_update_password_missing_fields(self):
        try:
            """Test updating password with missing required fields"""
            response = self.app.post('/update_password', data=json.dumps({
                'email': 'test@example.com'
                # Password is missing
            }), content_type='application/json')
            
            self.assertEqual(response.status_code, 400)
            self.write_to_excel('test_update_password_missing_fields','Success','')
        except Exception as e:
            self.write_to_excel('test_update_password_missing_fields','Failure',str(e))
        # Further assertions can be made based on the specific error response

if __name__ == '__main__':
    unittest.main()
