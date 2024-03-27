import unittest
from unittest.mock import MagicMock, patch
from datetime import datetime
import sys
import os
from openpyxl import Workbook, load_workbook

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)

from app.services.omegaTron.fyers_service import get_fyers_quote


class TestSuccessfulQuoteRetrieval(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        try:
            cls.workbook = load_workbook("fyers_orders_test.xlsx")
        except FileNotFoundError:
            cls.workbook = Workbook()

        cls.sheet = cls.workbook.active
        cls.sheet.title = "Login Results"
        if cls.sheet.max_row == 1:
            cls.sheet.append(
                [
                    "test case",
                    "result",
                    "notes",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(cls):
        cls.workbook.save("fyers_orders_test.xlsx")

    def write_to_excel(self, test_case, test_result, notes):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, notes, date, time])

    @patch("app.services.omegaTron.fyers_service.fyers_get_strike_price_ltp")
    def test_successful_quote_retrieval(self, mock_fyers_get_strike_price_ltp):
        mock_ce_option_chain = {"NIFTY2420820750CE": {"ltp": 1250}}
        mock_pe_option_chain = {"NIFTY2420820750PE": {"ltp": 0.15}}

        mock_fyers_get_strike_price_ltp.side_effect = [
            mock_ce_option_chain,
            mock_pe_option_chain,
        ]

        try:
            ce_option_chain, pe_option_chain, symbol = get_fyers_quote()
            try:
                self.assertDictEqual(ce_option_chain, mock_ce_option_chain)
            except AssertionError as e:
                print("Assertion Error in ce_option_chain:", e)
                self.write_to_excel("test_successful_quote_retrieval", "Failed", str(e))
            try:
                self.assertDictEqual(pe_option_chain, mock_pe_option_chain)
            except AssertionError as e:
                print("Assertion Error in pe_option_chain:", e)
                self.write_to_excel("test_successful_quote_retrieval", "Failed", str(e))
            self.assertIsInstance(ce_option_chain, dict)
            self.assertIsInstance(pe_option_chain, dict)
            self.assertEqual(symbol, "finnifty")
            self.write_to_excel(
                "test_successful_quote_retrieval", "Passes", "Test case passed"
            )
            print(
                "ce_option_chain ::: ",
                ce_option_chain,
                "pe_option_chain ::: ",
                pe_option_chain,
                "symbol ::: ",
                symbol,
            )
        except Exception as e:
            print("Exception occurred:", e)
            self.write_to_excel("test_successful_quote_retrieval", "Failed", str(e))

    #! A test case with the negetive values
    @patch("app.services.omegaTron.fyers_service.fyers_get_strike_price_ltp")
    def test_with_negetive_values(self, mock_fyers_get_strike_price_ltp):
        mock_ce_option_chain = {"NIFTY2420820750CE": {"ltp": -1250}}
        mock_pe_option_chain = {"NIFTY2420820750PE": {"ltp": -0.15}}

        mock_fyers_get_strike_price_ltp.side_effect = [
            mock_ce_option_chain,
            mock_pe_option_chain,
        ]

        try:
            ce_option_chain, pe_option_chain, symbol = get_fyers_quote()
            try:
                self.assertDictEqual(ce_option_chain, mock_ce_option_chain)
            except AssertionError as e:
                print("Assertion Error in ce_option_chain:", e)
                self.write_to_excel("test_with_negetive_values", "Failed", str(e))
            try:
                self.assertDictEqual(pe_option_chain, mock_pe_option_chain)
            except AssertionError as e:
                print("Assertion Error in pe_option_chain:", e)
                self.write_to_excel("test_with_negetive_values", "Failed", str(e))
            self.assertIsInstance(ce_option_chain, dict)
            self.assertIsInstance(pe_option_chain, dict)
            self.assertEqual(symbol, "midcpnifty")
            self.write_to_excel(
                "test_with_negetive_values", "Passes", "Test case passed"
            )
            print(
                "ce_option_chain ::: ",
                ce_option_chain,
                "pe_option_chain ::: ",
                pe_option_chain,
                "symbol ::: ",
                symbol,
            )
        except Exception as e:
            print("Exception occurred:", e)
            self.write_to_excel("test_with_negetive_values", "Failed", str(e))

    #! A test case with wrong symbol
    @patch("app.services.omegaTron.fyers_service.fyers_get_strike_price_ltp")
    def test_with_wrong_symbol(self, mock_fyers_get_strike_price_ltp):
        mock_ce_option_chain = {"NIFTY2420820750CE": {"ltp": 1250}}
        mock_pe_option_chain = {"NIFTY2420820750PE": {"ltp": 0.15}}

        mock_fyers_get_strike_price_ltp.side_effect = [
            mock_ce_option_chain,
            mock_pe_option_chain,
        ]

        try:
            ce_option_chain, pe_option_chain, symbol = get_fyers_quote()
            try:
                self.assertDictEqual(ce_option_chain, mock_ce_option_chain)
            except AssertionError as e:
                print("Assertion Error in ce_option_chain:", e)
                self.write_to_excel("test_with_wrong_symbol", "Failed", str(e))
            try:
                self.assertDictEqual(pe_option_chain, mock_pe_option_chain)
            except AssertionError as e:
                print("Assertion Error in pe_option_chain:", e)
                self.write_to_excel("test_with_wrong_symbol", "Failed", str(e))
            self.assertIsInstance(ce_option_chain, dict)
            self.assertIsInstance(pe_option_chain, dict)
            self.assertEqual(symbol, "test")
            self.write_to_excel("test_with_wrong_symbol", "Passes", "Test case passed")
            print(
                "ce_option_chain ::: ",
                ce_option_chain,
                "pe_option_chain ::: ",
                pe_option_chain,
                "symbol ::: ",
                symbol,
            )
        except Exception as e:
            print("Exception occurred:", e)
            self.write_to_excel("test_with_wrong_symbol", "Failed", str(e))


if __name__ == "__main__":
    unittest.main()
