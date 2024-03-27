import unittest
import json
from unittest.mock import patch
import os
import sys
from openpyxl import Workbook, load_workbook
from datetime import datetime

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)
from app import app

class TestLoginAPI(unittest.TestCase):

    @classmethod
    def setUpClass(self):
        self.app = app.test_client()
        self.app.testing = True
        try:
            self.workbook = load_workbook("login.xlsx")
        except FileNotFoundError:
            self.workbook = Workbook()

        self.sheet = self.workbook.active
        self.sheet.title = "Test result"
        if self.sheet.max_row == 1:
            self.sheet.append(
                [
                    "test case",
                    "result",
                    "notes",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(self):
        self.workbook.save("login.xlsx")

    def write_to_excel(self, test_case, test_result, notes):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, notes, date, time])

    def test_invalid_login(self):
        try:
            with patch('app.routes.userModule.login.execute_query') as mock_get_user:
                # Mock database response for failed authentication
                mock_get_user.return_value = None
                
                response = self.app.post('/login', data=json.dumps({
                    'email': 'wrong@example.com',
                    'password': 'wrong_password'
                }), content_type='application/json')
                
                self.assertEqual(response.status_code, 401)
                self.assertIn('Invalid email or password', json.loads(response.data)['message'])
                self.write_to_excel('test_invalid_login','Success','')
        except Exception as e:
            self.write_to_excel('test_invalid_login','Failure',str(e))

    def test_login_exception(self):
        try:
            with patch('app.routes.userModule.login.execute_query', side_effect=Exception('Test exception')):
                response = self.app.post('/login', data=json.dumps({
                    'email': 'test@example.com',
                    'password': 'whatever'
                }), content_type='application/json')
                
                self.assertEqual(response.status_code, 400)
                self.assertIn('error', json.loads(response.data))
                self.assertEqual('Test exception', json.loads(response.data)['error'])
                self.write_to_excel('test_login_exception','Success','')
        except Exception as e:
            self.write_to_excel('test_login_exception','Failure',str(e))
            

if __name__ == '__main__':
    unittest.main()
