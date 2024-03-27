import unittest
from unittest.mock import patch
from flask import json
import os, sys
from openpyxl import Workbook, load_workbook
from datetime import datetime

project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
sys.path.append(project_root)
from app import app

class SignupTestCase(unittest.TestCase):

    @classmethod
    def setUpClass(self):
        self.app = app.test_client()
        self.app.testing = True
        try:
            self.workbook = load_workbook("signup_test.xlsx")
        except FileNotFoundError:
            self.workbook = Workbook()

        self.sheet = self.workbook.active
        self.sheet.title = "Test result"
        if self.sheet.max_row == 1:
            self.sheet.append(
                [
                    "test case",
                    "result",
                    "notes",
                    "Date",
                    "Time",
                ]
            )

    @classmethod
    def tearDownClass(self):
        self.workbook.save("signup_test.xlsx")

    def write_to_excel(self, test_case, test_result, notes):
        date = datetime.now().strftime("%Y-%m-%d")
        time = datetime.now().strftime("%H:%M:%S")
        self.sheet.append([test_case, test_result, notes, date, time])


    @patch('app.routes.userModule.sign_up.execute_query')
    @patch('app.routes.userModule.sign_up.execute_query')
    def test_successful_signup(self, mock_add_user, mock_check_for_existing_email):
        try:
            """Test successful user signup"""
            mock_check_for_existing_email.return_value = []
            mock_add_user.return_value = True

            response = self.app.post('/signup', data=json.dumps({
                'name': 'Test User',
                'email': 'test@example.com',
                'password': 'testpassword',
                'dob' : '1990-05-15'
            }), content_type='application/json')
            
            self.assertEqual(response.status_code, 200)
            self.assertEqual(json.loads(response.data)['message'], 'User created successfully')
            self.write_to_excel('test_successful_signup','Success','')
        except Exception as e:
            self.write_to_excel('test_successful_signup','Failure',str(e))

    @patch('app.routes.userModule.sign_up.execute_query')
    def test_signup_with_existing_email(self, mock_check_for_existing_email):
        try:
            """Test signup with an existing email"""
            mock_check_for_existing_email.return_value = [True]

            response = self.app.post('/signup', data=json.dumps({
                'name': 'Test User',
                'email': 'existing@example.com',
                'password': 'testpassword'
            }), content_type='application/json')
            
            self.assertEqual(response.status_code, 409)
            self.assertEqual(json.loads(response.data)['message'], 'Email id already exists')
            self.write_to_excel('test_signup_with_existing_email','Success','')
        except Exception as e:
            self.write_to_excel('test_signup_with_existing_email','Failure',str(e))

    def test_signup_missing_fields(self):
        try:
            """Test signup with missing required fields"""
            response = self.app.post('/signup', data=json.dumps({
                'name': 'Test User',
                'password': 'testpassword'
                # Email is missing
            }), content_type='application/json')
            
            # Assuming validate_details returns a 400 for missing fields
            self.assertEqual(response.status_code, 400)
            # Further assertions can be made based on the specific error response
            self.write_to_excel('test_signup_missing_fields','Success','')
        except Exception as e:
            self.write_to_excel('test_signup_missing_fields','Failure',str(e))

if __name__ == '__main__':
    unittest.main()
